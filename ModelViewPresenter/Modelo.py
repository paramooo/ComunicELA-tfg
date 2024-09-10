from Servicios.Detector import Detector
from Adapter.Camara import Camara
from pygame.mixer import init as mixerInit, Sound as mixerSound
from random import randint
from numpy import zeros as np_zeros, zeros_like as np_zeros_like, uint8 as np_uint8, array as np_array, \
    ndarray as np_ndarray, expand_dims as np_expand_dims, concatenate as np_concatenate, min as np_min, \
    max as np_max, clip as np_clip, sqrt as np_sqrt, pi as np_pi, median as np_median, mean as np_mean, \
    polyfit as np_polyfit, poly1d as np_poly1d, sum as np_sum, unique as np_unique, where as np_where, \
    delete as np_delete, arctan2 as np_arctan2, cos as np_cos, sin as np_sin, pi as np_pi
from os import getenv, path as os_path, makedirs as os_makedirs
from kivy.app import App
from KivyCustom.Mensajes import Mensajes
from cv2 import addWeighted as cv2_addWeighted, \
    circle as cv2_circle, line as cv2_line, copyMakeBorder as cv2_copyMakeBorder, resize as cv2_resize, \
    BORDER_REPLICATE as cv2_BORDER_REPLICATE, INTER_AREA as cv2_INTER_AREA, imwrite as cv2_imwrite
from torch import from_numpy as torch_from_numpy, load as torch_load
from threading import Thread as threading_Thread
from torch import nn
import torch.optim as optim
from kivy.clock import Clock
from json import load as json_load, dump as json_dump
from PIL import Image, ImageDraw, ImageFont
from torch.nn.functional import mse_loss
from optuna import create_study
from openpyxl import load_workbook
from google.api_core.exceptions import RetryError
from google.api_core.retry import Retry
from google.generativeai import configure as genai_configure, GenerativeModel as genai_GenerativeModel
from win32com.client import Dispatch
from socket import create_connection
from scipy.ndimage import gaussian_filter1d
from ajustes.utils import get_recurso

class Modelo:
    """
    Clase Modelo del patrón MVP que se encarga de la lógica de la aplicación
    
    """
    def __init__(self):
        # Se inicializan las diferentes clases
        self.detector = Detector()
        self.camara = Camara()
        self.camara_act = None
        self.desarrollador = False
        self.camaras = self.obtener_camaras()

        # Crear el archivo de configuración de la aplicación para guardar los datos
        app_data_dir = getenv('LOCALAPPDATA')
        self.config_path = os_path.join(app_data_dir, 'ComunicELA', 'config.json')
        os_makedirs(os_path.dirname(self.config_path), exist_ok=True)
        if not os_path.exists(self.config_path):
            config = {
                "mostrar_tutorial": True,
                "idioma": "gal_ES",
                "corrector_frases": True,
                "camara": None,
                "voz": None,
            }
            with open(self.config_path, 'w') as config_file:
                json_dump(config, config_file)

        # Iniciacion de gemini
        api_key = getenv('GOOGLE_API_KEY')
        self.modelo_gemini = None
        if api_key is not None:
            genai_configure(api_key=api_key)
            self.modelo_gemini = genai_GenerativeModel('gemini-1.5-flash')
            
        # Inicializar el sintetizador de voz
        self.text_to_speech =  Dispatch("SAPI.SpVoice")

        # Para los sonidos
        mixerInit(buffer=4096)

        #Cargamos el config con la camara, corrector, idioma y voz
        with open(self.config_path, 'r') as f:
            config = json_load(f)
            if config["camara"] in self.camaras:
                self.iniciar_camara(config["camara"])
                self.camara_act = config["camara"]
            else:
                self.camara_act = None
            self.corrector_frases = config["corrector_frases"]
            #Si no es la primera vez que se abre
            voz_description = config["voz"]
            if voz_description != None:
                for voz in self.text_to_speech.GetVoices():
                    if voz.GetDescription() == voz_description:
                        self.text_to_speech.Voice = voz
                        break

        # Cargar el archivo de idioma correspondiente
        with open(get_recurso(f"strings/{self.get_idioma()}.json"), "r", encoding='utf-8') as f:
            self.strings = json_load(f)
        self.tamaño_fuente_txts = 25

        # Variables para el modelo y el postprocesado
        self.modelo_org = get_recurso('Componentes/aprox1_9Final.pt')
        self.postprocs = True
        self.modelo = torch_load(self.modelo_org)


        # Variables de control para la calibracion del parpadeo
        self.estado_calibracion = 0
        self.umbral_ear = 0.2
        self.umbral_ear_bajo = 0.2
        self.umbral_ear_cerrado = 0.2
        self.contador_p = 0
        self.suma_frames = 5 #Numero de frames que tiene que estar cerrado el ojo para que se considere un parpadeo
        self.calibrado = False
        self.sonido_click = mixerSound(get_recurso('sonidos/click.wav'))


        # Variables para la recopilacion de datos 
        self.reiniciar_datos_r()
        self.reiniciar_datos_reent()
        self.input = []
        self.output = []
        self.input_frames = []

        # Variable para el modelo
        self.pos_t = (0, 0)
        self.escanear = False

        # Variables para suavizar la mirada en el test
        self.historial = []     
        self.historial2 = []
        self.cantidad_suavizado = 30    # Suaviza la mirada con las medianas esto baja el ruido
        self.cantidad_suavizado2 = 5    # Suaviza las medianas asi el puntero se mueve suave
        self.hist_max = 60
        #self.retroceso_click = 0

        # Variables para uso de los tableros
        self.tableros = {}
        self.frase = ""
        self.tablero = None
        self.bloqueado = False
        self.sonido_alarma = mixerSound(get_recurso('sonidos/alarm.wav'))
        self.sonido_lock = mixerSound(get_recurso('sonidos/lock.wav'))
        self.pictogramas = False
        
        #variables para las pruebas de la aplicacion
        self.cronometro_pruebas = 0 
        self.contador_borrar = 0

        #Ponderación del 5% de la pantalla
        self.limiteAbajoIzq_org = [0.05,0.05]
        self.limiteAbajoDer_org = [0.95,0.05]
        self.limiteArribaIzq_org = [0,0.95]
        self.limiteArribaDer_org = [0.95,0.95]
        self.Desplazamiento_org = [0.5,0.5]

        # Variables para la ponderacion
        self.limiteAbajoIzq = self.limiteAbajoIzq_org
        self.limiteAbajoDer = self.limiteAbajoDer_org
        self.limiteArribaIzq = self.limiteArribaIzq_org
        self.limiteArribaDer = self.limiteArribaDer_org
        self.Desplazamiento = self.Desplazamiento_org

        # Aplicar la mascara a la imagen para la silueta
        self.fondo_frame_editado = Image.open(get_recurso('imagenes/fondo_marco_amp.png')).convert('L')
        self.mask_rgb = np_zeros((*np_array(self.fondo_frame_editado).shape, 3), dtype=np_uint8)
        self.mask_rgb[np_array(self.fondo_frame_editado)<50] = [50, 50, 50]

        #Variables para el reentrenamiento
        self.numero_entrenamientos = 0

        #Variables para la optimizacion
        self.optimizar = True #Variable para acivar o desactivar el optimizador al acabar el reentreno
        self.optimizando = False
        self.progreso_opt = 0
        self.trials_opt = 70

        

    
# ---------------------------   FUNCIONES DE CONTROL GENERAL    -------------------------------
#-------------------------------------------------------------------------------------------

    def reiniciar_datos_r(self):
        """
        Función que reinicia los datos de la recopilación de datos despues de cada escaneo, se aprovecha para inicializarlos

        """
        self.recopilar = False #Variable para saber si se esta recopilando datos
        self.contador_r = 5 #Contador para la cuenta atras
        self.pos_r = (0, 0) #Posicion de la pelota roja
        self.salto_bajo, self.salto_alto = 60, 80 #Salto de la pelota roja
        self.velocidad = 30
        self.direccion = 1 



    def reiniciar_datos_reent(self):
        """
        Función que reinicia los datos de la recopilación de datos despues de cada reentrenamiento, se aprovecha para inicializarlos
        Se recogen menos datos que en la recopilación de datos ya que esto es para solamente reajustar al usuario

        """
        self.recopilarRe = False
        self.salto_bajo_re, self.salto_alto_re = 100, 180
        self.velocidad_re = 35
        self.numero_epochs = 40
        self.porcentaje_reent = 0
        


    def datos_as_array(self, datos):
        """
        Función que transforma los datos en un array para poder ser utilizados por el modelo

        Args:
            datos (list): Lista con los datos de entrada
        
        Returns:
            np.ndarray: Array con los datos transformados

        """
        distancias_izq, distancias_der, or_x, or_y, or_z, ear, umbral_ear, coord_cab, _ = datos
        datos_transformados = np_expand_dims(np_concatenate([distancias_izq, distancias_der, [or_x], [or_y], [or_z], coord_cab, [ear], [umbral_ear]]), axis=0)
        return datos_transformados



    def conjunto_1(self, data):
        """
        Pre-procesado inicial, el cual ha dado mejores resultados

        Args:
            data (np.ndarray): Datos a pre-procesar
        
        Returns:
            np.ndarray: Datos pre-procesados

        Comentarios:
        Entradas: 39 -> TODOS LOS DATOS (NORMALIZAR MIN MAX DISTANCIAS)
        [0-15] Distancias entre los puntos de referencia ojo derecho min max
        [16-31] Distancias entre los puntos de referencia ojo izquierdo min max
        [32-34] Coordenadas de la orientación de la cara
        [35-36] Coordenadas del centro de la cara
        [37-38] EAR y umbral EAR 
        """
        # Normalizar cada valor de los primeros 16 de cada fila entre ellos mismos PARA NORMALIZAR EL OJO DERECHO
        data[:, :16] = (data[:, :16] - np_min(data[:, :16], axis=1, keepdims=True)) / (np_max(data[:, :16], axis=1, keepdims=True) - np_min(data[:, :16], axis=1, keepdims=True))

        # Normalizar cada valor de los segundos 16 de cada fila entre ellos mismos PARA NORMALIZAR EL OJO IZQUIERDO
        data[:, 16:32] = (data[:, 16:32] - np_min(data[:, 16:32], axis=1, keepdims=True)) / (np_max(data[:, 16:32], axis=1, keepdims=True) - np_min(data[:, 16:32], axis=1, keepdims=True))

        return data



    def get_string(self, clave):
        """
        Función que devuelve el string correspondiente a la clave en el idioma actual

        Args:
            clave (str): Clave del string a devolver
        
        Returns:
            str: String correspondiente a la clave en el idioma actual

        """
        return self.strings[clave]



    def get_show_tutorial(self):    
        """
        Función que devuelve si se debe mostrar el tutorial

        """
        with open(self.config_path, 'r') as f:
            config = json_load(f)
            return config["mostrar_tutorial"]

        
    def set_show_tutorial(self, valor):
        """
        Función que cambia el valor de mostrar tutorial

        Args:
            valor (bool): Valor a cambiar

        """
        with open(self.config_path, 'r') as f:
            config = json_load(f)
        config["mostrar_tutorial"] = valor
        with open(self.config_path, 'w') as f:
            json_dump(config, f)



    def cambiar_idioma(self):
        """
        Función que cambia el idioma de la aplicación

        """

        idioma = "gal_ES" if self.get_idioma() == "es_ES" else "es_ES"
        with open(self.config_path, 'r') as f:
            config = json_load(f)
        config["idioma"] = idioma
        with open(self.config_path, 'w') as f:
            json_dump(config, f)
        with open(get_recurso(f"strings/{idioma}.json"), "r", encoding='utf-8') as f:
            self.strings = json_load(f)

        
        

    def get_idioma(self):
        """
        Función que devuelve el idioma actual de la aplicación

        """
        try:
            with open(self.config_path, 'r') as f:
                config = json_load(f)
                return config["idioma"]
        except FileNotFoundError:
            return "es_ES"
        


    def get_idioma_string(self):
        """
        Función que devuelve el idioma actual de la aplicación en string

        """
        idiomas = {
            "es_ES": "Español",
            "gal_ES": "Galego",
        }
        return idiomas.get(self.get_idioma(), "Galego")



    def get_idioma_imagen(self):
        """
        Devuelve la imagen del idioma actual

        """
        return get_recurso(f'imagenes/idiomas/{self.get_idioma()}.png')
    


    def get_corrector_frases(self, mensajes = True):
        """
        Devuelve el estado del corrector de frases

        Args:
            mensajes (bool): Si se deben mostrar mensajes de error

        Returns:
            bool: Estado del corrector de frases

        """
        if self.modelo_gemini == None:
            if mensajes:
                self.mensaje(self.get_string('mensaje_api'))
            return None
        elif self.internet_available() == False:
            if mensajes:
                self.mensaje(self.get_string('mensaje_internet'))
            return None
        else:
            return self.corrector_frases
    


    def cambiar_estado_corrector(self):
        """
        Cambia el estado del conjugador automatico de frases

        """
        estado = self.get_corrector_frases()
        if estado is not None:
            self.corrector_frases = not self.corrector_frases
            estado = self.corrector_frases
            with open(self.config_path, 'r') as f:
                config = json_load(f)
            config["corrector_frases"] = self.corrector_frases
            with open(self.config_path, 'w') as f:
                json_dump(config, f)
        return estado



    def api_bien_configurada(self):
        """
        Función que devuelve si la API de Google está bien configurada

        """
        return self.modelo_gemini is not None
    

        
    def mensaje(self, mensaje):
        """
        Lanza un mensaje en la pantalla actual 

        Args:
            mensaje (str): Mensaje a mostrar
        
        """
        App.get_running_app().root.current_screen.add_widget(Mensajes(mensaje))



    def tarea_hilo(self, funcion):
        """
        Lanza una tarea en un hilo separado

        """
        # Crear un hilo para la tarea
        hilo = threading_Thread(target=funcion)
        hilo.start()



    def salir(self):
        """
        Cierra la aplicación

        """
        App.get_running_app().stop()





# ---------------------------   FUNCIONES DE CONTROL DE LA CAMARA    -------------------------------
#-------------------------------------------------------------------------------------------

    def iniciar_camara(self, index):
        """
        Inicia la cámara con el índice especificado

        """
        self.camara.start(index)



    def detener_camara(self):
        """
        Detiene la cámara actual

        """
        # Se detiene el escaneo de los ojos
        self.camara.stop()



    def camara_activa(self):
        """
        Devuelve si hay una cámara activa

        """
        return self.camara.camara_activa()
    


    def get_frame(self):
        """
        Devuelve el frame actual de la cámara
            
        """
        return self.camara.get_frame()
    


    def obtener_camaras(self, stop = True):
        """
        Devuelve las cámaras disponibles

        """
        return self.camara.obtener_camaras(stop)
    
    def seleccionar_camara(self, camara):
        """
        Activa la cámara con el índice especificado

        Args:
            camara (int): Índice de la cámara a activar
        """
        
        if self.camara_act != camara:
            if self.camara_activa():
                self.detener_camara()
            if camara is not None:
                self.iniciar_camara(camara)
            self.camara_act = camara
            #giardar en el config
            with open(self.config_path, 'r') as f:
                config = json_load(f)
            config["camara"] = camara
            with open(self.config_path, 'w') as f:
                json_dump(config, f)



    def get_index_actual(self):
        """
        Devuelve el indice de la cámara activa

        """
        return self.camara_act



    def get_frame_editado(self):
        """
        Devuelve el frame editado con la silueta de la cara y los puntos pintados

        """
        frame = self.get_frame()
        if frame is None:
            # Croger la imagen con la forma
            frame_pil = Image.fromarray(self.mask_rgb)

            # Seleccionar la fuente y el tamaño
            font = ImageFont.truetype(get_recurso("KivyCustom/fuentes/FrancoisOne-Regular.ttf"), 30)

            # Calcular el ancho del texto
            text = self.get_string('mensaje_frame_editado')
            
            # Dibujar el texto en la imagen
            draw = ImageDraw.Draw(frame_pil)
            draw.text((180, 430), text, font=font, fill=(255, 255, 255, 0))

            # Convertir la imagen de PIL de vuelta a un array de numpy
            frame = np_array(frame_pil)
            
            return frame

        # Blanco por defecto
        color = (255, 255, 255)  
        
        # Coger los puntos
        coord_central = None
        puntos_or = None
        datos = self.detector.obtener_coordenadas_indices(frame)
        if datos is not None:
            _, _, _, _, coord_central, puntos_or = datos
        frame = cv2_addWeighted(frame, 0.5, self.mask_rgb, 1, 0)

        # Crear un circulo en el medio del frame
        r = 10
        if coord_central is not None:
            x = round(frame.shape[1]*coord_central[0])
            y = round(frame.shape[0]*coord_central[1])

            dx = x - frame.shape[1]//2
            dy = y - frame.shape[0]//2

            # Si el punto central esta dentro del circulo de radio r pintar de verde
            if dx**2 + dy**2 < r**2:
                color = (0, 255, 0)
            else:
                color = (255, 255, 255)

                # Calcular el ángulo de la línea desde el centro del círculo hasta el punto central
                angle = np_arctan2(dy, dx)

                # Calcular las coordenadas del punto en el borde del círculo
                x_end = frame.shape[1]//2 + r * np_cos(angle)
                y_end = frame.shape[0]//2 + r * np_sin(angle)

                # Dibujar la línea principal de la flecha
                cv2_line(frame, (x, y), (int(x_end), int(y_end)), color, 2)

                # Calcular las coordenadas de los puntos de la punta de la flecha
                arrow_size = np_clip(np_sqrt(dx**2 + dy**2) / 5, 1, 100)
                dx_arrow1 = arrow_size * np_cos(angle + np_pi/4)  
                dy_arrow1 = arrow_size * np_sin(angle + np_pi/4)
                dx_arrow2 = arrow_size * np_cos(angle - np_pi/4)  
                dy_arrow2 = arrow_size * np_sin(angle - np_pi/4)

                # Dibujar las líneas de la punta de la flecha
                cv2_line(frame, (int(x_end), int(y_end)), (int(x_end + dx_arrow1), int(y_end + dy_arrow1)), color, 2)
                cv2_line(frame, (int(x_end), int(y_end)), (int(x_end + dx_arrow2), int(y_end + dy_arrow2)), color, 2)

            # Crear un circulo de radio r en el centro y un punto en el centro de la cara
            cv2_circle(frame, (frame.shape[1]//2, frame.shape[0]//2), r, color, 2)
            cv2_circle(frame, (round(frame.shape[1]*coord_central[0]), round(frame.shape[0]*coord_central[1])), 5, color, -1)   

        # Colorear los otros puntos de la cara
        if puntos_or is not None:
            for punto in puntos_or:
                x = punto[0]
                y = punto[1]
                frame[y - 1:y + 1, x - 1:x + 1, :] = color


        return frame




# --------------------------- FUNCIONES PARA EL CONTROL DE LAS VOCES -------------------------
#-------------------------------------------------------------------------------------------

    def get_voces(self):
        """
        Devuelve las voces disponibles instaladas en el sistema

        """
        voces = self.text_to_speech.GetVoices()
        voces_description = []
        for i in range(len(voces)):
            voces_description.append(voces[i].GetDescription())
        return voces_description



    def seleccionar_voz(self, voz_description):
        """
        Selecciona la voz mediante la descripción de la misma

        Args:
            voz_description (str): Descripción de la voz a seleccionar
        
        """
        for voz in self.text_to_speech.GetVoices():
            if voz.GetDescription() == voz_description:
                self.text_to_speech.Voice = voz
                #Apuntar la voz en el config
                with open(self.config_path, 'r') as f:
                    config = json_load(f)
                config["voz"] = voz_description
                with open(self.config_path, 'w') as f:
                    json_dump(config, f)
                break



    def get_voz_seleccionada(self):
        """
        Devuelve la voz seleccionada actualmente

        """
        return self.text_to_speech.Voice.GetDescription()


    
# ---------------------------   FUNCIONES CONTROL DEL MENU DE CALIBRACION -------------------------------
#------------------------------------------------------------------------------------------
                

    def obtener_estado_calibracion(self):
        """
        Devuelve el estado actual de la calibración (fase1, fase2, fase3)

        """

        return self.estado_calibracion        
    


    def cambiar_estado_calibracion(self, numero = None):
        """
        Actualiza el estado de la calibración obteniendo los datos necesarios en el paso de cada fase

        Args:
            numero (int): Número de la fase de la calibración

        Returns:
            int: Número de la fase de la calibración

        """
        if numero is not None:
            self.estado_calibracion = numero
            return self.estado_calibracion
        # Se obtienen los datos
        datos = self.detector.obtener_coordenadas_indices(self.get_frame())

        #Si no se detecta cara, None
        if datos is None:
            self.mensaje(self.get_string("mensaje_calibracion_fallida"))
            return None
    
        # Se desempaquetan los datos
        _, _, coord_ear_izq, coord_ear_der, _, _ = self.detector.obtener_coordenadas_indices(self.get_frame())

        # Se captura el EAR actual dependiendo del estado de la calibracion
        if self.estado_calibracion == 1:
            self.umbral_ear_bajo = self.detector.calcular_ear_medio(coord_ear_izq, coord_ear_der)

        elif self.estado_calibracion == 2:
            self.umbral_ear_cerrado = self.detector.calcular_ear_medio(coord_ear_izq, coord_ear_der)
            self.umbral_ear = (self.umbral_ear_bajo*0.4 + self.umbral_ear_cerrado*0.6) #Se calcula el umbral final ponderado entre el cerrado y el abierto bajo
            self.calibrado = True
        
        elif self.estado_calibracion == 3:
            self.estado_calibracion = 0
            return self.estado_calibracion

        self.estado_calibracion += 1
        return self.estado_calibracion



    def get_parpadeo(self, ear):
        """
        Obtiene si se ha producido un parpadeo

        Args:
            ear (float): Valor del EAR actual
        
        Returns:
            int: 1 si se ha producido un parpadeo, 0 en caso contrario
        """
        if ear < self.umbral_ear:
            # Contador para parpadeo
            self.contador_p += 1

            #Si se mantiene cerrado el ojo durante 60 frames, se bloquea el tablero
            if self.contador_p == 60:  
                self.bloqueado = not self.bloqueado
                self.sonido_lock.play()
            
            #Si se mantiene cerrado el ojo durante suma_frames, se considera un parpadeo
            if self.contador_p == self.suma_frames:
                if not self.bloqueado:
                    self.sonido_click.play()
                return 1
            return 0
        else:
            self.contador_p = 0
            return 0   
    



    def set_limite(self, valor, esquina, eje):
        """
        Establece los valores de las ponderaciones a realizar

        Args:
            valor (float): Valor de la ponderación
            esquina (int): Esquina a la que pertenece la ponderación
            eje (int): Eje al que pertenece la ponderación
        """
        if esquina == 0:
            self.limiteAbajoIzq[eje] = valor
        elif esquina == 1:
            self.limiteAbajoDer[eje] = valor
        elif esquina == 2:
            self.limiteArribaIzq[eje] = valor
        elif esquina == 3:
            self.limiteArribaDer[eje] = valor
        elif esquina == 4:
            self.Desplazamiento[eje] = valor



    def get_limites(self):
        """
        Devuelve los valores de las ponderaciones
            
        """
        return self.limiteAbajoIzq, self.limiteAbajoDer, self.limiteArribaIzq, self.limiteArribaDer, self.Desplazamiento





# ---------------------------   FUNCIONES DE RECOPILACION DE DATOS  -------------------------------
#--------------------------------------------------------------------------------------------------
        
    def cuenta_atras(self, dt):
        """
        Función que realiza la cuenta atrás para la recopilación de datos

        """
        if self.contador_r > 0:
            self.contador_r -= 1
        elif self.contador_r == 0:
            Clock.unschedule(self.cuenta_atras)
            return False



    def actualizar_pos_circle_r(self, tamano_pantalla):
        """
        Función que actualiza la posición de la pelota roja en la pantalla

        Args:
            tamano_pantalla (tuple): Tamaño de la pantalla
        
        Returns:
            tuple: Posición de la pelota roja
        """
        velocidad = self.velocidad_re if self.recopilarRe else self.velocidad
        salto_bajo = self.salto_bajo_re if self.recopilarRe else self.salto_bajo
        salto_alto = self.salto_alto_re if self.recopilarRe else self.salto_alto
    
        # Actualiza la posición x de la pelota
        self.pos_r = (self.pos_r[0] + velocidad * self.direccion, self.pos_r[1])

        # Si la pelota toca los bordes x de la pantalla, cambia la dirección y realiza un salto
        if self.pos_r[0] < 0 or self.pos_r[0] + 50 > tamano_pantalla[0]:
            self.direccion *= -1

            # Actualiza la posición y de la pelota con un salto aleatorio
            salto = randint(min(salto_bajo, salto_alto), max(salto_bajo, salto_alto))
            self.pos_r = (self.pos_r[0], self.pos_r[1] + salto)

        # Si la pelota toca el borde superior de la pantalla, invierte el salto
        if self.pos_r[1] + 50 > tamano_pantalla[1]: 
            # Invertimos los saltos y bajamos un poco
            self.salto_bajo, self.salto_alto , self.salto_bajo_re, self.salto_alto_re = self.salto_bajo*-1, self.salto_alto*-1, self.salto_bajo_re*-1, self.salto_alto_re*-1
            self.pos_r = (self.pos_r[0], tamano_pantalla[1] - 50)

        # Si la pelota toca el borde inferior de la pantalla
        if self.pos_r[1] <= 0 and self.salto_alto < 0:
            # Si viene de el reentrenamiento, se reentrena 
            if self.recopilarRe:
                self.tarea_hilo(lambda: self.reentrenar())
                self.reiniciar_datos_reent()
            else:
                self.reiniciar_datos_r()
        else:
            datos = self.obtener_datos()
            if datos is not None:                
                distancias_izq, distancias_der, or_x, or_y, or_z, ear, umbral_ear, coord_cab, frame = datos
                self.guardar_datos(distancias_izq, distancias_der, or_x, or_y, or_z, ear, umbral_ear, coord_cab, self.pos_r/np_array(tamano_pantalla), frame)
        return self.pos_r



    def guardar_datos(self, distancias_izq, distancias_der, or_x, or_y, or_z, ear, umbral_ear, coord_cab, pos_r_norm, frame):
        """
        Guarda los datos obtenidos en las listas de datos

        Args:
            distancias_izq (list): Distancias de los puntos de referencia del ojo izquierdo
            distancias_der (list): Distancias de los puntos de referencia del ojo derecho
            or_x (float): Coordenada x de la orientación de la cara
            or_y (float): Coordenada y de la orientación de la cara
            or_z (float): Coordenada z de la orientación de la cara
            ear (float): Valor del EAR
            umbral_ear (float): Valor del umbral del EAR
            coord_cab (list): Coordenadas del centro de la cara
            pos_r_norm (tuple): Posición de la pelota roja normalizada
            frame (np.ndarray): Frame actual

        """
        # Guardar los datos en las listas
        self.input.append([*distancias_izq, *distancias_der, or_x, or_y, or_z, *coord_cab, ear, umbral_ear])
        self.output.append(pos_r_norm)
        self.input_frames.append(frame)



    def guardar_final(self):
        """
        Guarda los datos de las listas en los diferentes archivos en segundo plano

        """
        def guardar_aux():
            # Determinar el número de líneas existentes en el archivo
            with open(f'./entrenamiento/datos/txts/input.txt', 'r') as f:
                num_lineas = sum(1 for _ in f)+1

            # Guardar los datos en los archivos
            with open(f'./entrenamiento/datos/txts/input.txt', 'a') as f:
                for i, linea in enumerate(self.input):
                    # Convertir el elemento a cadena si es una lista o tupla
                    if isinstance(linea, (list, tuple, np_ndarray)):
                        linea = ', '.join(map(str, linea))
                    f.write(str(linea) + '\n')
                    cv2_imwrite(f'./entrenamiento/datos/frames/frame_{num_lineas}.jpg', self.input_frames[i])
                    num_lineas += 1

            with open(f'./entrenamiento/datos/txts/output.txt', 'a') as f:
                for linea in self.output:
                    # Convertir el elemento a cadena si es una lista o tupla
                    if isinstance(linea, (list, tuple, np_ndarray)):
                        linea = ', '.join(map(str, linea))
                    f.write(str(linea) + '\n')

            # Limpiar las listas para la próxima vez
            self.input = []
            self.output = []
            self.input_frames = []

            print('Datos guardados correctamente')
        self.tarea_hilo(lambda: guardar_aux())

    
    def descartar_datos(self):
        """
        Descarta los datos recopilados por si se ha producido un error al recopilarlos

        """
        self.input = []
        self.output = []
        self.input_frames = []


       


# --------------------------- FUNCIONES DE OBTENCION DE LA POSICIÓN DE LA MIRADA Y LOS DATOS  ----------
#-------------------------------------------------------------------------------------------------------

    def obtener_datos(self):
        """
        Obtiene los diferentes datos necesarios del frame actual

        """
        frame = self.get_frame()
        if frame is None:
            self.mensaje(self.get_string("mensaje_error_camara"))
            return None
        datos = self.detector.obtener_coordenadas_indices(frame)
        
        #Si no se detecta cara, se devuelve None
        if datos is None:
            self.mensaje(self.get_string("mensaje_no_cara"))
            return None
        
        # Se desempaquetan los datos
        coord_o_izq, coord_o_der, coord_ear_izq, coord_ear_der, coord_cab, coord_o = datos

        # Distancias de los ojos 
        distancias_izq, distancias_der = self.detector.calcular_distancias_ojos(coord_o_izq, coord_o_der)

        # Orientación de la cabeza entre 0 y 1
        or_x, or_y, or_z = self.detector.get_orientacion_cabeza(coord_o, frame.shape)

        # EAR 
        ear = self.detector.calcular_ear_medio(coord_ear_izq, coord_ear_der)

        # Pasamos la posicion de la pantalla normalizada
        return distancias_izq, distancias_der, or_x, or_y, or_z, ear, self.umbral_ear, coord_cab, frame



    def obtener_posicion_mirada_ear(self):
        """
        Devuelve la posición de la mirada y si se ha producido un click
        """
        # Se obtienen los datos
        datos = self.obtener_datos()

        # Si no se detecta cara
        if datos is None:
            return None

        # Se desempaquetan los datos del ear para el click
        _, _, _, _, _, ear, _, _, _ = datos
        click = self.get_parpadeo(ear)

        # Se transforman los datos 
        datos_array = self.datos_as_array(datos)
        datos_array = self.conjunto_1(datos_array)
        datos_array = torch_from_numpy(datos_array).float()

        # Se predice la posición de la mirada
        mirada = self.modelo(datos_array)

        # Se desempaqueta la posición de la mirada
        mirada = mirada.data.numpy()[0]

        # Postprocesar la posición de la mirada
        if self.postprocs:
            mirada = self.postprocesar(mirada)

        return mirada, click



    def postprocesar(self, mirada):
        """
        Se postprocesa la posición resultante del modelo, se suaviza y se pondera

        Args:
            mirada (np.ndarray): Posición de la mirada
        
        Returns:
            np.ndarray: Posición de la mirada postprocesada
        """
        # Ponderar la mirada
        mirada = self.ponderar(mirada)

        # Añadir la nueva posición al historial
        self.historial.append(mirada)

        # Eliminar la posición más asntigua si el historial es demasiado largo
        if len(self.historial) > self.hist_max:
            self.historial.pop(0)

        # Primero con la mediana para eliminar ruido y no perder tanto retraso
        mirada = np_median(self.historial[-self.cantidad_suavizado:], axis=0)

        self.historial2.append(mirada)
        if len(self.historial2) > self.hist_max:
            self.historial2.pop(0)
        # Despues con la media de las medianas para suavizar el trazado del puntero
        mirada = np_mean(self.historial2[-self.cantidad_suavizado2:], axis=0)
        
        return mirada



    def ponderar(self, mirada, limiteAbajoIzq=None, limiteAbajoDer=None, limiteArribaIzq=None, limiteArribaDer=None, Desplazamiento=None):
        """
        Pondera la posición de la mirada en función de los límites de la pantalla establecidos

        Args:
            mirada (np.ndarray): Posición de la mirada
            limiteAbajoIzq (np.ndarray): Límite inferior izquierdo de la pantalla
            limiteAbajoDer (np.ndarray): Límite inferior derecho de la pantalla
            limiteArribaIzq (np.ndarray): Límite superior izquierdo de la pantalla
            limiteArribaDer (np.ndarray): Límite superior derecho de la pantalla
            Desplazamiento (np.ndarray): Desplazamiento de la mirada
        
        Returns:
            np.ndarray: Posición de la mirada ponderada
        """

        def calcular_limites_esquina(cuadrante):
            """
            Funcion auxiliar que devuelve los limites de la esquina en función de la esquina que sea (tabla de explicativa en la memoria)

            Args:   
                cuadrante (int): Cuadrante de la esquina

            Returns:
                tuple: Limites de la esquina
            """
            if limiteAbajoDer is None:
                if cuadrante == 0:
                    return self.limiteAbajoIzq[0], self.limiteAbajoIzq[1], self.limiteAbajoDer[0], self.limiteArribaIzq[1]
                elif cuadrante == 1:
                    return self.limiteAbajoIzq[0], self.limiteAbajoDer[1], self.limiteAbajoDer[0], self.limiteArribaDer[1]
                elif cuadrante == 2:
                    return self.limiteArribaIzq[0], self.limiteAbajoIzq[1], self.limiteArribaDer[0], self.limiteArribaIzq[1]
                elif cuadrante == 3:
                    return self.limiteArribaIzq[0], self.limiteAbajoDer[1], self.limiteArribaDer[0], self.limiteArribaDer[1]
            else:
                if cuadrante == 0:
                    return limiteAbajoIzq[0], limiteAbajoIzq[1], limiteAbajoDer[0], limiteArribaIzq[1]
                elif cuadrante == 1:
                    return limiteAbajoIzq[0], limiteAbajoDer[1], limiteAbajoDer[0], limiteArribaDer[1]
                elif cuadrante == 2:
                    return limiteArribaIzq[0], limiteAbajoIzq[1], limiteArribaDer[0], limiteArribaIzq[1]
                elif cuadrante == 3:
                    return limiteArribaIzq[0], limiteAbajoDer[1], limiteArribaDer[0], limiteArribaDer[1]
        

        def ponderar_esquina(mirada, esquina_limites):
            """
            Funcion auxiliar que pondera para una esquina en concreto

            Args:
                mirada (np.ndarray): Posición de la mirada
                esquina_limites (tuple): Limites de la esquina
            
            Returns:
                np.ndarray: Posición de la mirada ponderada
            """
            LimiteBajoX, LimiteBajoY, LimiteAltoX, LimiteAltoY = esquina_limites

            # Calculamos los límites de la zona no afectada
            ComienzoZonaNoAfectadaX = LimiteBajoX + ((self.Desplazamiento[0] if limiteAbajoDer is None else Desplazamiento[0]) - LimiteBajoX) / 2
            FinZonaNoAfectadaX = LimiteAltoX - (LimiteAltoX - (self.Desplazamiento[0] if limiteAbajoDer is None else Desplazamiento[0])) / 2
            ComienzoZonaNoAfectadaY = LimiteBajoY + ((self.Desplazamiento[1] if limiteAbajoDer is None else Desplazamiento[1]) - LimiteBajoY) / 2
            FinZonaNoAfectadaY = LimiteAltoY - (LimiteAltoY - (self.Desplazamiento[1] if limiteAbajoDer is None else Desplazamiento[1])) / 2

            # Calculamos las x y las y de las Xs
            Xx = np_array([LimiteBajoX, ComienzoZonaNoAfectadaX, (self.Desplazamiento[0] if limiteAbajoDer is None else Desplazamiento[0]), FinZonaNoAfectadaX, LimiteAltoX])
            Xy = np_array([0, ComienzoZonaNoAfectadaX, 0.5, FinZonaNoAfectadaX, 1])
            Yx = np_array([LimiteBajoY, ComienzoZonaNoAfectadaY, (self.Desplazamiento[1] if limiteAbajoDer is None else Desplazamiento[1]), FinZonaNoAfectadaY, LimiteAltoY])
            Yy = np_array([0, ComienzoZonaNoAfectadaY, 0.5, FinZonaNoAfectadaY, 1])

            # Crear la función polinómica
            polinomioX = np_poly1d(np_polyfit(Xx, Xy, 4))
            polinomioY = np_poly1d(np_polyfit(Yx, Yy, 4))

            # Calcular el valor ponderado
            return np_array([np_clip(polinomioX(mirada[0]), 0, 1), np_clip(polinomioY(mirada[1]), 0, 1)])


        def calcular_distancia(mirada, esquina):
            """
            Función auxiliar que calcula la distancia hacia la esquina para decidir el peso en la media ponderada de esta esquina

            Args:
                mirada (np.ndarray): Posición de la mirada
                esquina (np.ndarray): Esquina a la que se quiere calcular la distancia
            
            Returns:
                float: Distancia de la mirada a la esquina
            """
            return np_sqrt((mirada[0] - esquina[0])**2 + (mirada[1] - esquina[1])**2)


        # Definir las cuatro esquinas
        esquinas = []
        if limiteAbajoDer is None:
            esquinas = [self.limiteAbajoIzq, self.limiteAbajoDer, self.limiteArribaIzq, self.limiteArribaDer]
        else:
            esquinas = [limiteAbajoIzq, limiteAbajoDer, limiteArribaIzq, limiteArribaDer]
        ponderaciones = []

        # Calcular la ponderación para cada esquina
        for esquina_limites in esquinas:
            ponderacion_esquina = ponderar_esquina(mirada, calcular_limites_esquina(esquinas.index(esquina_limites)))
            ponderaciones.append(ponderacion_esquina)

        
        # Calcular la distancia de la mirada a cada esquina
        distancias = [calcular_distancia(mirada, esquina) for esquina in esquinas]

        # Normalizar las distancias para obtener pesos de ponderación
        pesos = np_array([1 / (distancia*2 + 1) for distancia in distancias])  # Cambio aquí

        # Realizar normalizacion min-max de los pesos
        pesos = (pesos - np_min(pesos)) / (np_max(pesos) - np_min(pesos))

        # Sumar los pesos
        suma_pesos = np_sum(pesos)

        # Ponderar las ponderaciones de acuerdo a las distancias
        ponderacion_final = np_zeros_like(ponderaciones[0])
        for i, ponderacion_esquina in enumerate(ponderaciones):
            ponderacion_final += ponderacion_esquina * (pesos[i] / suma_pesos)

        return ponderacion_final

            



#--------------------------------FUNCIONES PARA LOS TABLEROS--------------------------------
#-------------------------------------------------------------------------------------------
    def cargar_tableros(self):
        """
        Carga los tableros de palabras con imágenes en función del idioma seleccionado

        """
        try:
            idioma = self.get_idioma()
            filename = get_recurso(f'tableros/tableros_{idioma}.xlsx')
            if os_path.isfile(filename):
                wb = load_workbook(filename)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    palabras_con_imagenes = []
                    for row in ws.iter_rows():
                        fila = []
                        for i in range(0, len(row), 2):
                            imagen_celda = row[i]
                            palabra_celda = row[i + 1]
                            fila.append((imagen_celda.value, palabra_celda.value))
                        palabras_con_imagenes.append(fila)
                    self.tableros[sheet_name] = palabras_con_imagenes
                return True
            else:
                self.mensaje(self.get_string("mensaje_error_tableros"))
                return False
        except:
            self.mensaje(self.get_string("mensaje_error_tableros"))
            return False



    def obtener_tablero(self, nombre):
        """
        Devuelve el tablero correspondiente

        """
        return self.tableros.get(nombre)
    


    def obtener_tablero_inicial(self):
        """
        Devuelve el tabler inicial, el del indice 0
        """
        return list(self.tableros.keys())[0]
    

    
    def añadir_palabra(self, palabra):
        """
        Añade una palabra a la frase

        """
        self.frase += palabra + ' '



    def borrar_palabra(self):
        """
        Borra la última palabra de la frase

        """
        self.frase = ' '.join(self.frase.rstrip().split(' ')[:-1]) + ' '
        self.contador_borrar += 1
        # Verificar si la frase contiene solo espacios
        if self.frase.strip() == '':
            self.frase = ''



    def borrar_todo(self):
        """
        Borra todas las palabras de la frase

        """
        numero_palabras = len(self.frase.split(' '))
        self.contador_borrar += numero_palabras
        self.frase = ''



    def alarma(self):
        """
        Reproduce el sonido de alarma

        """
        self.sonido_alarma.play()



    def internet_available(self):
        """
        Comprueba si hay conexión a internet abriendo una conexión con el servidor de Google

        """
        try:
            create_connection(("8.8.8.8", 53))
            return True
        except OSError:
            return False



    def get_frase_bien(self):
        """
        LLama a la API de Gemini para conjugar la frase y devolverla correctamente

        """
        frase = self.frase
        frase_mod = None

        if self.modelo_gemini is not None:
            prompt = ("Recibes una frase con palabras en infinitivo y el idioma en el que está escrita(Español o Gallego)." +
                    "Tu tarea es conjugar la frase para que las palabras estén en la forma correcta y coherente entre sí siendo coherente también con el idioma.\n" +
                    "No se permiten signos de puntuación, solo palabras y el simbolo de interrogación(esto solamente en caso de que se añada en el input).\n"+
                    "Devuelve SOLAMENTE la frase conjugada en la forma más previsible posible que crees que significa lo que quiera decir.\n"+
                    "Si la frase tiene solamente una palabra, puedes agregar un determinante si es necesario.\n"+
                    "Ejemlo 1 (es): Entrada: TÚ COMER CARNE?\nRespuesta: ¿Tú comes carne?"+
                    "Ejemlo 2 (es): Entrada: YO QUERER COMER CARNE\nRespuesta: Yo quiero comer carne"+
                    "Ejemlo 3 (gal): Entrada: EU NECESITAR VER DOUTOR\nRespuesta: Eu necesito ver ao doutor"+
                    "Ejmplo 4 (gal): Entrada: EU QUERER TI MOITO\nRespuesta: Eu quérote moito"+
                    "\n\nFrase: " + frase + "\nIdioma: " + self.get_idioma())
            if self.internet_available():
                try:
                    frase_mod = self.modelo_gemini.generate_content(prompt, request_options={'timeout': 5, 'retry': Retry(initial=1, multiplier=2, maximum=1, timeout=5)}, generation_config={'temperature': 0.1})
                except RetryError as e:
                    pass
            else:
                pass

            try:
                return frase_mod.text
            except ValueError:
                return frase
        else:
            return frase



    def reproducir_texto(self):
        """
        Emplea el Text-To-Speech de Windows para reproducir la frase

        """
        #Empezar un hulo separado:
        def reproducir_texto_hilo():
            if self.corrector_frases:
                self.frase = self.get_frase_bien().upper()
            if self.text_to_speech is not None:
                self.text_to_speech.Speak(self.frase)
            else:
                self.mensaje(self.get_string("mensaje_error_sintesis"))

        #Se crea un hilo para reproducir el texto
        if self.frase != '':
            self.tarea_hilo(lambda: reproducir_texto_hilo())
        else :
            self.mensaje(self.get_string("mensaje_frase_vacia"))


#--------------------------------FUNCIONES PARA LAS PRUEBAS--------------------------------
#-------------------------------------------------------------------------------------------

    def cronometro(self, dt):
        """
        Se encarga de llevar el cronometro de las pruebas
        """
        self.cronometro_pruebas += dt



    def iniciar_cronometro(self):
        """
        Inicia el cronometro de las pruebas en un hilo separado

        """
        Clock.schedule_interval(self.cronometro, 0.01)
        self.contador_borrar = 0



    def stop_cronometro(self):
        """
        Detiene el cronometro de las pruebas

        """
        Clock.unschedule(self.cronometro)



    def reiniciar_cronometro(self):
        """
        Reinicia el cronometro de las pruebas y el contador de borrado
        """
        Clock.unschedule(self.cronometro)
        self.cronometro_pruebas = 0
        self.contador_borrar = 0


    
    #----------------------------------- FUNCIONES PARA EL REEENTRENAMIENTO --------------------------------
    #-----------------------------------------------------------------------------------------------------

    def reentrenar(self):
        """
        Se encarga de reentrenar el modelo con los datos recopilados del usuario
        """

        # Se obtienen los datos
        self.input = np_array(self.input)
        self.output = np_array(self.output)

        # Si los datos estan vacios no se puede reentrenar
        if len(self.input) < 20:
            print("No hay datos para reentrenar")
            self.porcentaje_reent = -1
            self.input = []
            self.output = []
            return
        
        personas = []
        contador = 1
        for i in range(len(self.output)):
            if i == 0 or self.output[i][1] != 0.0 or self.output[i-1][1] == 0.0:
                personas.append(contador)
            else:
                contador += 1
                personas.append(contador)

        unique_persons = np_unique(personas)
        input = self.input

        for person in unique_persons:
            indices = np_where(personas == person)[0]
            for i in range(self.input.shape[1]-2):
                input[indices, i] = gaussian_filter1d(self.input[indices, i], 5)

        # Se eliminan los datos con el ojo cerrado
        index = np_where(self.input[:, -2] < self.input[:, -1])
        input = np_delete(self.input, index, axis=0)
        output = np_delete(self.output, index, axis=0)

        # Se obtiene el conjunto 
        input_conj = self.conjunto_1(input)

        # Se convierten a tensores
        input_train = torch_from_numpy(input_conj).float()
        output_train = torch_from_numpy(output).float()

        # Se reentrena al modelo
        optimizer = optim.Adam(self.modelo.parameters(), lr=0.00001)

        train_losses = []
        best_loss = float('inf')
        models = []
        loss = nn.MSELoss()
        first_loss = loss(self.modelo(input_train), output_train).item()
        
        if self.optimizar:
            self.optimizando = True
            self.progreso_opt = 0

        # print("Precisión antes del reentrenamiento: ", first_loss) 
        for epoch in range(self.numero_epochs+1):
            optimizer.zero_grad()

            # Entrenamiento y cálculo de la pérdida
            train_predictions = self.modelo(input_train)
            train_loss = loss(train_predictions, output_train)
            train_losses.append(train_loss.item())

            # Guardar el mejor modelo
            if train_loss.item() < best_loss:
                best_loss = train_loss.item()
            models.append(self.modelo)

            # Actualizar el modelo
            train_loss.backward()
            optimizer.step()

            #Guardar el porcentaje de epoch que llevamos
            self.porcentaje_reent = int((epoch/self.numero_epochs)*100)

        self.modelo = models[train_losses.index(min(train_losses))]
        

        # print("--------PERDIDAS DEL REENTRENAMIENTO----------")
        # print("Perdida final con reentrenamiento: ", min(train_losses))
        # print("Perdida antes del reentrenamiento: ", first_loss)

        # Se optimizan las esquinas
        if self.optimizar:
            error_optimizado = self.optimizar_esquinas(input, output)
            # print("Error con optimización: ", error_optimizado)



    def descartar_reentrenamientos(self):
        """
        Reestablece el modelo a su estado original y descarta los datos recopilados para el reentrenamiento

        """
        if self.numero_entrenamientos == 0:
            self.mensaje(self.get_string("mensaje_no_reentrenamientos"))
            return
        self.modelo = torch_load(self.modelo_org)
        self.mensaje(self.get_string("mensaje_descartados"))
        self.numero_entrenamientos = 0

        # Limpiamos los datos de reentrenamiento
        self.input = []
        self.output = []

        # Reajustamos las esquinas a su valor orignial
        self.limiteAbajoIzq = self.limiteAbajoIzq_org
        self.limiteAbajoDer = self.limiteAbajoDer_org
        self.limiteArribaIzq = self.limiteArribaIzq_org
        self.limiteArribaDer = self.limiteArribaDer_org
        self.Desplazamiento = self.Desplazamiento_org



    def optimizar_esquinas(self, input_opt, output_opt):
        """
        Realiza la busqueda de los mejores valores para las esquinas y el desplazamiento

        Args:
            input_opt (np.ndarray): Datos de entrada para la optimización
            output_opt (np.ndarray): Datos de salida para la optimización

        Returns:
            float: Mejor error obtenido con la optimización
        """
        # Se obtiene el conjunto y los tensores
        input_conj_opt = self.conjunto_1(input_opt)
        input_conj_opt = torch_from_numpy(input_conj_opt).float()

        def objective(trial):
            """
            Funcion auxiliar para la busqueda con Optuna
            """
            # Definir los límites de las esquinas y el desplazamiento
            limiteAbajoIzqX = trial.suggest_float('limiteAbajoIzqX', self.limiteAbajoIzq_org[0]-0.025, self.limiteAbajoIzq_org[0]+0.025)
            limiteAbajoIzqY = trial.suggest_float('limiteAbajoIzqY', self.limiteAbajoIzq_org[1]-0.025, self.limiteAbajoIzq_org[1]+0.025)
            limiteAbajoDerX = trial.suggest_float('limiteAbajoDerX', self.limiteAbajoDer_org[0]-0.025, self.limiteAbajoDer_org[0]+0.025)
            limiteAbajoDerY = trial.suggest_float('limiteAbajoDerY', self.limiteAbajoDer_org[1]-0.025, self.limiteAbajoDer_org[1]+0.025)
            limiteArribaIzqX = trial.suggest_float('limiteArribaIzqX', self.limiteArribaIzq_org[0]-0.025, self.limiteArribaIzq_org[0]+0.025)
            limiteArribaIzqY = trial.suggest_float('limiteArribaIzqY', self.limiteArribaIzq_org[1]-0.025, self.limiteArribaIzq_org[1]+0.025)
            limiteArribaDerX = trial.suggest_float('limiteArribaDerX', self.limiteArribaDer_org[0]-0.025, self.limiteArribaDer_org[0]+0.025)
            limiteArribaDerY = trial.suggest_float('limiteArribaDerY', self.limiteArribaDer_org[1]-0.025, self.limiteArribaDer_org[1]+0.025)
            DesplazamientoX = trial.suggest_float('DesplazamientoX', self.Desplazamiento_org[0]-0.025, self.Desplazamiento_org[0]+0.025)
            DesplazamientoY = trial.suggest_float('DesplazamientoY', self.Desplazamiento_org[1]-0.025, self.Desplazamiento_org[1]+0.025)

            limiteAbajoIzq = [limiteAbajoIzqX, limiteAbajoIzqY]
            limiteAbajoDer = [limiteAbajoDerX, limiteAbajoDerY]
            limiteArribaIzq = [limiteArribaIzqX, limiteArribaIzqY]
            limiteArribaDer = [limiteArribaDerX, limiteArribaDerY]
            Desplazamiento = [DesplazamientoX, DesplazamientoY]

            # Calcular las predicciones del modelo
            predicciones = self.modelo((input_conj_opt)).detach().numpy()

            for i, prediccion in enumerate(predicciones):
                predicciones[i] = self.ponderar(prediccion, limiteAbajoIzq, limiteAbajoDer, limiteArribaIzq, limiteArribaDer, Desplazamiento)

            
            # Calcular el error
            output_tensor = torch_from_numpy(output_opt).float()
            predicciones_tensor = torch_from_numpy(predicciones).float()

            error = mse_loss(output_tensor, predicciones_tensor).item()

            return error
        
        
        def actualizar_progreso_opt(study, trial):
            """
            Función auxiliar para actualizar el progreso de la optimización

            Args:
                study (optuna.study.Study): Estudio de Optuna
                trial (optuna.trial.Trial): Prueba actual
            """
            progreso = int(((trial.number+1) / self.trials_opt) * 100)
            if self.progreso_opt < progreso:
                self.progreso_opt = progreso

        # De donde parte la busqueda
        initial_params = {
        'limiteAbajoIzqX': self.limiteAbajoIzq_org[0],
        'limiteAbajoIzqY': self.limiteAbajoIzq_org[1],
        'limiteAbajoDerX': self.limiteAbajoDer_org[0],
        'limiteAbajoDerY': self.limiteAbajoDer_org[1],
        'limiteArribaIzqX': self.limiteArribaIzq_org[0],
        'limiteArribaIzqY': self.limiteArribaIzq_org[1],
        'limiteArribaDerX': self.limiteArribaDer_org[0],
        'limiteArribaDerY': self.limiteArribaDer_org[1],
        'DesplazamientoX': self.Desplazamiento_org[0],
        'DesplazamientoY': self.Desplazamiento_org[1]
        }

        study = create_study(direction='minimize')
        study.enqueue_trial(initial_params)
        study.optimize(objective, n_trials=self.trials_opt, callbacks=[actualizar_progreso_opt],show_progress_bar=False, n_jobs=3)

        # Guardar los resultados
        self.limiteAbajoIzq = [study.best_params['limiteAbajoIzqX'], study.best_params['limiteAbajoIzqY']]
        self.limiteAbajoDer = [study.best_params['limiteAbajoDerX'], study.best_params['limiteAbajoDerY']]
        self.limiteArribaIzq = [study.best_params['limiteArribaIzqX'], study.best_params['limiteArribaIzqY']]
        self.limiteArribaDer = [study.best_params['limiteArribaDerX'], study.best_params['limiteArribaDerY']]
        self.Desplazamiento = [study.best_params['DesplazamientoX'], study.best_params['DesplazamientoY']]
        
        self.optimizando = False
        self.progreso_opt = 0
        return study.best_value  