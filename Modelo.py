from Detector import Detector
from Camara import Camara
from pygame import mixer
import random
import numpy as np
import os
from kivy.app import App
from Mensajes import Mensajes
from entrenamiento.Conjuntos import Conjuntos
import cv2
import torch
import threading
import math
from torch import nn
import torch.optim as optim
from kivy.clock import Clock
import json
from PIL import Image, ImageDraw, ImageFont
from torch.nn.functional import mse_loss
import optuna
import openpyxl
import google.generativeai as genai
import win32com.client

class Modelo:
    def __init__(self):
        # Para el sonido del click
        mixer.init(buffer=4096)

        # Se inicializa el detector
        self.detector = Detector()
        self.camara = Camara()
        self.camara_act = None
        self.desarrollador = False
        # Iniciacion de gemini
        api_key = os.getenv('GOOGLE_API_KEY')
        self.modelo_gemini = None
        if api_key is not None:
            genai.configure(api_key=api_key)
            self.modelo_gemini = genai.GenerativeModel('gemini-1.5-flash')
        
        # Cargar el archivo de idioma correspondiente
        with open(f"./strings/{self.get_idioma()}.json", "r", encoding='utf-8') as f:
            self.strings = json.load(f)

        # Variables para el modelo de test
        self.conjunto = 1
        self.modelo_org = './entrenamiento/modelos/aprox1_9.pt'
        self.postprocs = False
        # self.conjunto = 2
        # self.modelo_org = './entrenamiento/modelos/modelo_ajustado.pth'
        
        self.modelo = torch.load(self.modelo_org)

        # Variables de control para el tamaño de la fuente de los textos
        self.tamaño_fuente_txts = 25

        # Variables de control para la calibracion del parpadeo
        self.estado_calibracion = 0
        self.umbral_ear = 0.2
        self.umbral_ear_bajo = 0.2
        self.umbral_ear_cerrado = 0.2
        self.contador_p = 0
        self.suma_frames = 4 #Numero de frames que tiene que estar cerrado el ojo para que se considere un parpadeo
        self.calibrado = False
        self.sonido_click = mixer.Sound('./sonidos/click.wav')


        # Variables para la recopilacion de datos 
        self.reiniciar_datos_r()
        self.reiniciar_datos_reent()
        self.input = []
        self.output = []
        self.input_frames = []

        # Variable para el modelo
        self.pos_t = (0, 0)
        self.escanear = False

        # Variables para suavizar la mirada en el test
        self.historial = []     # Suaviza la mirada con la mediana esto baja el ruido
        self.historial2 = []    # Suaviza las medianas asi el puntero se mueve suave
        self.cantidad_suavizado = 18
        self.cantidad_suavizado2 = 5
        self.hist_max = 90
        #self.retroceso_click = 0

        # Variables para uso de los tableros
        self.tableros = {}
        self.cargar_tableros()
        self.frase = ""
        self.tablero = None
        self.bloqueado = False
        self.contador_pb = 0
        self.sonido_alarma = mixer.Sound('./sonidos/alarm.wav')
        self.sonido_lock = mixer.Sound('./sonidos/lock.wav')
        self.pictogramas = False
        
        #variables para las pruebas de la aplicacion
        self.cronometro_pruebas = 0 #Variable para el cronometro de las pruebas
        self.contador_borrar = 0

        # Ponderar la mirada del ajustado
        # self.limiteAbajoIzq_org = [0.10,0.07]
        # self.limiteAbajoDer_org = [0.87,0.09]
        # self.limiteArribaIzq_org = [0.03,0.81]
        # self.limiteArribaDer_org = [0.93,0.89]
        # self.Desplazamiento_org = [0.5,0.5]

        #SIN PONDERAR 
        self.limiteAbajoIzq_org = [0,0]
        self.limiteAbajoDer_org = [1,0]
        self.limiteArribaIzq_org = [0,1]
        self.limiteArribaDer_org = [1,1]
        self.Desplazamiento_org = [0.5,0.5]

        # Variables para la ponderacion
        self.limiteAbajoIzq = self.limiteAbajoIzq_org
        self.limiteAbajoDer = self.limiteAbajoDer_org
        self.limiteArribaIzq = self.limiteArribaIzq_org
        self.limiteArribaDer = self.limiteArribaDer_org
        self.Desplazamiento = self.Desplazamiento_org

        # Aplicar un umbral
        self.fondo_frame_editado = cv2.imread('./imagenes/fondo_marco_amp.png', cv2.IMREAD_GRAYSCALE)
        self.mask_rgb = np.zeros((*self.fondo_frame_editado.shape, 3), dtype=np.uint8)
        self.mask_rgb[self.fondo_frame_editado<50] = [50, 50, 50]

        #Variables para el reentrenamiento
        self.numero_entrenamientos = 0

        #Variables para la optimizacion
        self.optimizando = False
        self.progreso_opt = 0
        self.optimizar = False #Variable para acivar o desactivar el optimizador al acabar el reentreno
        self.trials_opt = 70


    #Funcion para reiniciar los datos despues de cada escaneo (se aprovecha para inicializarlos tambien)
    def reiniciar_datos_r(self):
        self.recopilar = False #Variable para saber si se esta recopilando datos
        self.contador_r = 5 #Contador para la cuenta atras
        self.pos_r = (0, 0) #Posicion de la pelota roja
        self.salto_bajo, self.salto_alto = 60, 80 #Salto de la pelota roja
        self.velocidad = 25
        self.direccion = 1 

    #Funcion para reiniciar los datos despues de cada reentrenamiento (se aprovecha para inicializarlos tambien)
    #Son menos que en la recopilacion ya que esto es para solamente reajustar al usuario
    def reiniciar_datos_reent(self):
        self.recopilarRe = False
        self.salto_bajo_re, self.salto_alto_re = 100, 180
        self.velocidad_re = 35
        self.numero_epochs = 20
        self.porcentaje_reent = 0
        


    def get_string(self, clave):
        return self.strings[clave]

    # Cargamos la configuracion del tutorial
    def get_show_tutorial(self):    
        try:
            with open('./config.json', 'r') as f:
                config = json.load(f)
                return config["mostrar_tutorial"]
        except FileNotFoundError:
            return True  # Por defecto, mostrar el tutorial
        
    def set_show_tutorial(self, valor):
        config = {"mostrar_tutorial": valor}
        with open('config.json', 'w') as f:
            json.dump(config, f)

    def cambiar_idioma(self):
        idioma = "gal_ES" if self.get_idioma() == "es_ES" else "es_ES"
        with open('config.json', 'r') as f:
            config = json.load(f)
        config["idioma"] = idioma
        with open('config.json', 'w') as f:
            json.dump(config, f)

        #Actualizar el idioma de los strings
        with open(f"./strings/{idioma}.json", "r", encoding='utf-8') as f:
            self.strings = json.load(f)
        #Actualizar el idioma de los tableros
        self.cargar_tableros()
        
        

    def get_idioma(self):
        try:
            with open('./config.json', 'r') as f:
                config = json.load(f)
                return config["idioma"]
        except FileNotFoundError:
            return "es_ES"
        
    def get_idioma_string(self):
        idiomas = {
            "es_ES": "Español",
            "gal_ES": "Galego",
        }
        return idiomas.get(self.get_idioma(), "Galego")

    def get_idioma_imagen(self):
        return f'./imagenes/idiomas/{self.get_idioma()}.png'
    
    
    
# ---------------------------   FUNCIONES DE CONTROL GENERAL    -------------------------------
#-------------------------------------------------------------------------------------------
        
    def mensaje(self, mensaje):
        # Crear el mensaje de mensaje
        mensaje = Mensajes(mensaje)

        # Agregar el mensaje a la pantalla actual
        App.get_running_app().root.current_screen.add_widget(mensaje)


    def tarea_hilo(self, funcion):
        # Crear un hilo para la tarea
        hilo = threading.Thread(target=funcion)
        hilo.start()

    def salir(self):
        App.get_running_app().stop()

# ---------------------------   FUNCIONES DE CONTROL DE LA CAMARA    -------------------------------
#-------------------------------------------------------------------------------------------

    def iniciar_camara(self, index):
        # Se inicia el escaneo de los ojos
        self.camara.start(index)

    def detener_camara(self):
        # Se detiene el escaneo de los ojos
        self.camara.stop()

    def camara_activa(self):
        return self.camara.camara_activa()
    
    def get_frame(self):
        return self.camara.get_frame()
    
    def obtener_camaras(self):
        return self.camara.obtener_camaras()
    
    def seleccionar_camara(self, camara):
        if self.camara_activa():
            self.detener_camara()
        self.iniciar_camara(camara)
        self.camara_act = camara

    def get_index_actual(self):
        return self.camara_act


    def get_frame_editado(self):
        frame = self.get_frame()
        if frame is None:
            # Croger la imagen con la forma
            frame_pil = Image.fromarray(self.mask_rgb)

            # Seleccionar la fuente y el tamaño
            font = ImageFont.truetype("./kivy/FrancoisOne-Regular.ttf", 30)

            # Calcular el ancho del texto
            text = self.get_string('mensaje_frame_editado')
            
            # Dibujar el texto en la imagen
            draw = ImageDraw.Draw(frame_pil)
            draw.text((180, 430), text, font=font, fill=(255, 255, 255, 0))

            # Convertir la imagen de PIL de vuelta a un array de numpy
            frame = np.array(frame_pil)
            
            return frame

        # Blanco por defecto
        color = (255, 255, 255)  
        
        # Coger los puntos
        coord_central = self.detector.get_punto_central(frame)
        puntos_or = self.detector.get_puntos_or(frame)

        # Ahora puedes usar mask_rgb en lugar de self.mask

        frame = cv2.addWeighted(frame, 0.5, self.mask_rgb, 1, 0)

        # Crear un circulo en el medio del frame
        r = 10
        # Poner el punto central y la flecha en caso de necesitarla
        if coord_central is not None:
            x = round(frame.shape[1]*coord_central[0])
            y = round(frame.shape[0]*coord_central[1])

            dx = x - frame.shape[1]//2
            dy = y - frame.shape[0]//2

            # Si el punto central esta dentro del circulo de radio r
            if dx**2 + dy**2 < r**2:
                color = (0, 255, 0)
            else:
                color = (255, 255, 255)

                # Calcular el ángulo de la línea desde el centro del círculo hasta el punto central
                angle = math.atan2(dy, dx)

                # Calcular las coordenadas del punto en el borde del círculo
                x_end = frame.shape[1]//2 + r * math.cos(angle)
                y_end = frame.shape[0]//2 + r * math.sin(angle)

                # Dibujar la línea principal de la flecha
                cv2.line(frame, (x, y), (int(x_end), int(y_end)), color, 2)

                # Calcular las coordenadas de los puntos de la punta de la flecha
                arrow_size = np.clip(np.sqrt(dx**2 + dy**2) / 5, 1, 100)
                dx_arrow1 = arrow_size * math.cos(angle + np.pi/4)  # Ajusta el ángulo para la primera línea de la punta de la flecha
                dy_arrow1 = arrow_size * math.sin(angle + np.pi/4)
                dx_arrow2 = arrow_size * math.cos(angle - np.pi/4)  # Ajusta el ángulo para la segunda línea de la punta de la flecha
                dy_arrow2 = arrow_size * math.sin(angle - np.pi/4)

                # Dibujar las líneas de la punta de la flecha
                cv2.line(frame, (int(x_end), int(y_end)), (int(x_end + dx_arrow1), int(y_end + dy_arrow1)), color, 2)
                cv2.line(frame, (int(x_end), int(y_end)), (int(x_end + dx_arrow2), int(y_end + dy_arrow2)), color, 2)

            # Crear un circulo de radio r en el centro
            cv2.circle(frame, (frame.shape[1]//2, frame.shape[0]//2), r, color, 2)

            # Crear un circulo de radio 10 en el punto central
            cv2.circle(frame, (round(frame.shape[1]*coord_central[0]), round(frame.shape[0]*coord_central[1])), 5, color, -1)   

        # Colorear los puntos de la cara
        if puntos_or is not None:
            for punto in puntos_or:
                x = round(frame.shape[1]*punto[0])
                y = round(frame.shape[0]*punto[1])
                frame[y - 1:y + 1, x - 1:x + 1, :] = color


        return frame

# ---------------------------   FUNCIONES CONTROL DEL MENU DE CALIBRACION -------------------------------
#------------------------------------------------------------------------------------------
                

    def obtener_estado_calibracion(self):
        return self.estado_calibracion        
    

# ---------------------------   FUNCIONES DE CONTROL DEL EAR -------------------------------
#-------------------------------------------------------------------------------------------
    
#Para la calibracion
    def cambiar_estado_calibracion(self, numero = None):
        if numero is not None:
            self.estado_calibracion = numero
            return self.estado_calibracion
        # Se obtienen los datos
        datos = self.detector.obtener_coordenadas_indices(self.get_frame())

        #Si no se detecta cara, None
        if datos is None:
            self.mensaje(self.get_string("mensaje_calibracion_fallida"))
            return None
    
        # Se desempaquetan los datos
        _, _, coord_ear_izq, coord_ear_der, _, _ = self.detector.obtener_coordenadas_indices(self.get_frame())

        # Se captura el EAR actual dependiendo del estado de la calibracion
        if self.estado_calibracion == 1:
            self.umbral_ear_bajo = self.detector.calcular_ear_medio(coord_ear_izq, coord_ear_der)

        elif self.estado_calibracion == 2:
            self.umbral_ear_cerrado = self.detector.calcular_ear_medio(coord_ear_izq, coord_ear_der)
            self.umbral_ear = (self.umbral_ear_bajo*0.4 + self.umbral_ear_cerrado*0.6) #Se calcula el umbral final ponderado entre el cerrado y el abierto bajo
            self.calibrado = True
        
        elif self.estado_calibracion == 3:
            self.estado_calibracion = 0
            return self.estado_calibracion

        self.estado_calibracion += 1
        return self.estado_calibracion

#Para el test/tableros
    def get_parpadeo(self, ear):
        if ear < self.umbral_ear:
            # Contador para parpadeo
            self.contador_p += 1
            # Contador para el bloqueo de los tableros
            self.contador_pb += 1

            #Si se mantiene cerrado el ojo durante 60 frames, se bloquea el tablero
            if self.contador_pb == 60:  
                self.contador_pb = -1000
                self.bloqueado = not self.bloqueado
                self.sonido_lock.play()
            
            #Si se mantiene cerrado el ojo durante suma_frames, se considera un parpadeo
            if self.contador_p == self.suma_frames:
                self.contador_p = -1000 
                if not self.bloqueado:
                    self.sonido_click.play()
                return 1
            return 0
        else:
            self.contador_p = 0
            self.contador_pb = 0     
            return 0   
    
    def alarma(self):
        self.sonido_alarma.play()

    def set_limite(self, valor, esquina, eje):
        if esquina == 0:
            self.limiteAbajoIzq[eje] = valor
        elif esquina == 1:
            self.limiteAbajoDer[eje] = valor
        elif esquina == 2:
            self.limiteArribaIzq[eje] = valor
        elif esquina == 3:
            self.limiteArribaDer[eje] = valor
        elif esquina == 4:
            self.Desplazamiento[eje] = valor

    def get_limites(self):
        return self.limiteAbajoIzq, self.limiteAbajoDer, self.limiteArribaIzq, self.limiteArribaDer, self.Desplazamiento

# ---------------------------   FUNCIONES DE RECOPILACION DE DATOS  -------------------------------
        
    def cuenta_atras(self, dt):
        if self.contador_r > 0:
            self.contador_r -= 1
        elif self.contador_r == 0:
            Clock.unschedule(self.cuenta_atras)
            return False

    def actualizar_pos_circle_r(self, tamano_pantalla):
        velocidad = self.velocidad_re if self.recopilarRe else self.velocidad
        salto_bajo = self.salto_bajo_re if self.recopilarRe else self.salto_bajo
        salto_alto = self.salto_alto_re if self.recopilarRe else self.salto_alto
    
        # Actualiza la posición x de la pelota
        self.pos_r = (self.pos_r[0] + velocidad * self.direccion, self.pos_r[1])

        # Si la pelota toca los bordes x de la pantalla, cambia la dirección y realiza un salto
        if self.pos_r[0] < 0 or self.pos_r[0] + 50 > tamano_pantalla[0]:
            self.direccion *= -1

            # Actualiza la posición y de la pelota con un salto aleatorio
            salto = random.randint(min(salto_bajo, salto_alto), max(salto_bajo, salto_alto))
            self.pos_r = (self.pos_r[0], self.pos_r[1] + salto)

        # Si la pelota toca el borde superior de la pantalla, invierte el salto
        if self.pos_r[1] + 50 > tamano_pantalla[1]: 
            # Invertimos los saltos y bajamos un poco
            self.salto_bajo, self.salto_alto , self.salto_bajo_re, self.salto_alto_re = self.salto_bajo*-1, self.salto_alto*-1, self.salto_bajo_re*-1, self.salto_alto_re*-1
            self.pos_r = (self.pos_r[0], tamano_pantalla[1] - 50)

        # Si la pelota toca el borde inferior de la pantalla
        if self.pos_r[1] < 0:
            # Si viene de el reentrenamiento, se reentrena 
            if self.recopilarRe:
                self.tarea_hilo(lambda: self.reentrenar())
                self.reiniciar_datos_reent()
            else:
                self.reiniciar_datos_r()
        else:
            datos = self.obtener_datos()
            if datos is not None:                
                distancias_izq, distancias_der, or_x, or_y, or_z, ear, umbral_ear, coord_cab, frame = datos
                self.guardar_datos(distancias_izq, distancias_der, or_x, or_y, or_z, ear, umbral_ear, coord_cab, self.pos_r/np.array(tamano_pantalla), frame)
        return self.pos_r


    def guardar_datos(self, distancias_izq, distancias_der, or_x, or_y, or_z, ear, umbral_ear, coord_cab, pos_r_norm, frame):
        # Guardar los datos en las listas
        self.input.append([*distancias_izq, *distancias_der, or_x, or_y, or_z, *coord_cab, ear, umbral_ear])
        self.output.append(pos_r_norm)
        self.input_frames.append(frame)


    def guardar_final(self, fichero):
        def guardar_aux(fichero):
            # Determinar el número de líneas existentes en el archivo
            with open(f'./entrenamiento/datos/txts/input{fichero}.txt', 'r') as f:
                num_lineas = sum(1 for _ in f)+1

            # Guardar los datos en los archivos
            with open(f'./entrenamiento/datos/txts/input{fichero}.txt', 'a') as f:
                for i, linea in enumerate(self.input):
                    # Convertir el elemento a cadena si es una lista o tupla
                    if isinstance(linea, (list, tuple, np.ndarray)):
                        linea = ', '.join(map(str, linea))
                    f.write(str(linea) + '\n')
                    cv2.imwrite(f'./entrenamiento/datos/frames/{fichero}/frame_{num_lineas}.jpg', self.input_frames[i])
                    num_lineas += 1

            with open(f'./entrenamiento/datos/txts/output{fichero}.txt', 'a') as f:
                for linea in self.output:
                    # Convertir el elemento a cadena si es una lista o tupla
                    if isinstance(linea, (list, tuple, np.ndarray)):
                        linea = ', '.join(map(str, linea))
                    f.write(str(linea) + '\n')

            # Limpiar las listas para la próxima vez
            self.input = []
            self.output = []
            self.input_frames = []

            print('Datos guardados correctamente')
        self.tarea_hilo(lambda: guardar_aux(fichero))

    
    def descartar_datos(self):
        self.input = []
        self.output = []
        self.input_frames = []


       


# ---------------------------   FUNCIONES DE OBTENCION DE DATOS  -------------------------------
#-----------------------------------------------------------------------------------------------
    def obtener_datos(self):
        frame = self.get_frame()
        if frame is None:
            self.mensaje(self.get_string("mensaje_error_camara"))
            return None
        datos = self.detector.obtener_coordenadas_indices(frame)
        
        #Si no se detecta cara, se devuelve None
        if datos is None:
            self.mensaje(self.get_string("mensaje_no_cara"))
            return None
        
        # Se desempaquetan los datos
        coord_o_izq, coord_o_der, coord_ear_izq, coord_ear_der, coord_cab, coord_o = datos

        # - Distancias de los ojos 
        distancias_izq, distancias_der = self.detector.calcular_distancias_ojos(coord_o_izq, coord_o_der)

        # - Orientación de la cabeza entre 0 y 1
        or_x, or_y, or_z = self.detector.get_orientacion_cabeza(coord_o, frame.shape)

        # - EAR 
        ear = self.detector.calcular_ear_medio(coord_ear_izq, coord_ear_der)

        # - Recortar el rectangulo de los ojos normalizado a 200x50 ESTO SE HACE EN EL POSTPROCESAR NO AL OBTENER LOS DATOS
        #rect_frame = self.normalizar_frame(frame, coord_o_izq, coord_o_der)

        #print("ORX: ", round(or_x,3), "ORY: ", round(or_y,3), "ORZ: ", round(or_z,3), "coord_cab: ", coord_cab)

        # Pasamos la posicion de la pantalla normalizada
        return distancias_izq, distancias_der, or_x, or_y, or_z, ear, self.umbral_ear, coord_cab, frame

    # Recortar el rectangulo de los ojos normalizado a 200x50
    def normalizar_frame(self, frame, coord_o_izq, coord_o_der):
        # Coordenadas de los ojos
        x_o_izq, y_o_izq = coord_o_izq[0]
        x_o_der, y_o_der = coord_o_der[0]

        # Coordenadas del rectangulo
        x1 = min(x_o_izq, x_o_der)-10
        x2 = max(x_o_izq, x_o_der)+10
        y1 = min(y_o_izq, y_o_der)-10
        y2 = max(y_o_izq, y_o_der)+10

        # Recortar el rectangulo de los ojos
        rect_frame = frame[y1:y2, x1:x2]

        # Calcular la relación de aspecto deseada
        ratio = 200.0 / 50.0
        ratio_act = rect_frame.shape[1] / rect_frame.shape[0]

        # Calcular cuántos píxeles se deben agregar a cada lado
        if ratio_act < ratio:
            diff = int((rect_frame.shape[0] * ratio - rect_frame.shape[1]) / 2)
            rect_frame = cv2.copyMakeBorder(rect_frame, 0, 0, diff, diff, cv2.BORDER_REPLICATE)
        elif ratio_act > ratio:
            diff = int((rect_frame.shape[1] / ratio - rect_frame.shape[0]) / 2)
            rect_frame = cv2.copyMakeBorder(rect_frame, diff, diff, 0, 0, cv2.BORDER_REPLICATE)

        # Redimensionar a 200x50 manteniendo la relación de aspecto
        rect_frame = cv2.resize(rect_frame, (200, 50), interpolation = cv2.INTER_AREA)

        return rect_frame


    #Funcion para obtener la posicion de la mirada en el test
    #Se obtiene la posición de la mirada y si se ha hecho click
    #Se suaviza la posición de la mirada y si se hace click, se retrocede self.retroceso_click frames para evitar la desviación
    #el error esta aqui en como se le pasan los datos en el datos as array creo 
    def obtener_posicion_mirada_ear(self):
        # Se obtienen los datos
        datos = self.obtener_datos()

        # Si no se detecta cara, se devuelve None, None
        if datos is None:
            return None

        # Se desempaquetan los datos del ear para el click
        _, _, _, _, _, ear, _, _, _ = datos
        click = self.get_parpadeo(ear)

        # Se transforman los datos a un conjunto
        datos_array = Conjuntos.datos_as_array(datos)
        normalizar_funcion = getattr(Conjuntos, f'conjunto_{self.conjunto}')
        datos_array = normalizar_funcion(datos_array)


        # Normalizacion para la ResNet
        #----------------------------------
        # datos_array = np.reshape(datos_array, (-1, 1, 23))
        # datos_array = np.repeat(datos_array[:, :, np.newaxis], 3, axis=1)
        # self.modelo.eval()
        #----------------------------------


        datos_array = torch.from_numpy(datos_array).float()

        # Se predice la posición de la mirada
        mirada = self.modelo(datos_array)

        # Se desempaqueta la posición de la mirada
        mirada = mirada.data.numpy()[0]

        print(mirada)

        # Postprocesar la posición de la mirada
        if self.postprocs:
            mirada = self.postprocesar(mirada)

        return mirada, click


    def postprocesar(self, mirada):
        # Ponderar la mirada
        mirada = self.ponderar(mirada)

        # Añadir la nueva posición al historial
        self.historial.append(mirada)

        # Eliminar la posición más asntigua si el historial es demasiado largo
        if len(self.historial) > self.hist_max:
            self.historial.pop(0)

        # Primero con la mediana para eliminar ruido y no perder tanto retraso
        mirada = np.median(self.historial[-self.cantidad_suavizado:], axis=0)

        self.historial2.append(mirada)
        if len(self.historial2) > self.hist_max:
            self.historial2.pop(0)
        # Despues con la media de las medianas para suavizar el trazado del puntero
        mirada = np.mean(self.historial2[-self.cantidad_suavizado2:], axis=0)
        
        return mirada


    def ponderar(self, mirada, limiteAbajoIzq=None, limiteAbajoDer=None, limiteArribaIzq=None, limiteArribaDer=None, Desplazamiento=None):
        def calcular_limites_esquina(cuadrante):
            if limiteAbajoDer is None:
                if cuadrante == 0:
                    return self.limiteAbajoIzq[0], self.limiteAbajoIzq[1], self.limiteAbajoDer[0], self.limiteArribaIzq[1]
                elif cuadrante == 1:
                    return self.limiteAbajoIzq[0], self.limiteAbajoDer[1], self.limiteAbajoDer[0], self.limiteArribaDer[1]
                elif cuadrante == 2:
                    return self.limiteArribaIzq[0], self.limiteAbajoIzq[1], self.limiteArribaDer[0], self.limiteArribaIzq[1]
                elif cuadrante == 3:
                    return self.limiteArribaIzq[0], self.limiteAbajoDer[1], self.limiteArribaDer[0], self.limiteArribaDer[1]
            else:
                if cuadrante == 0:
                    return limiteAbajoIzq[0], limiteAbajoIzq[1], limiteAbajoDer[0], limiteArribaIzq[1]
                elif cuadrante == 1:
                    return limiteAbajoIzq[0], limiteAbajoDer[1], limiteAbajoDer[0], limiteArribaDer[1]
                elif cuadrante == 2:
                    return limiteArribaIzq[0], limiteAbajoIzq[1], limiteArribaDer[0], limiteArribaIzq[1]
                elif cuadrante == 3:
                    return limiteArribaIzq[0], limiteAbajoDer[1], limiteArribaDer[0], limiteArribaDer[1]
        
        def ponderar_esquina(mirada, esquina_limites):
            LimiteBajoX, LimiteBajoY, LimiteAltoX, LimiteAltoY = esquina_limites

            # Calculamos los límites de la zona no afectada
            ComienzoZonaNoAfectadaX = LimiteBajoX + ((self.Desplazamiento[0] if limiteAbajoDer is None else Desplazamiento[0]) - LimiteBajoX) / 2
            FinZonaNoAfectadaX = LimiteAltoX - (LimiteAltoX - (self.Desplazamiento[0] if limiteAbajoDer is None else Desplazamiento[0])) / 2
            ComienzoZonaNoAfectadaY = LimiteBajoY + ((self.Desplazamiento[1] if limiteAbajoDer is None else Desplazamiento[1]) - LimiteBajoY) / 2
            FinZonaNoAfectadaY = LimiteAltoY - (LimiteAltoY - (self.Desplazamiento[1] if limiteAbajoDer is None else Desplazamiento[1])) / 2

            # Calculamos las x y las y de las Xs
            Xx = np.array([LimiteBajoX, ComienzoZonaNoAfectadaX, (self.Desplazamiento[0] if limiteAbajoDer is None else Desplazamiento[0]), FinZonaNoAfectadaX, LimiteAltoX])
            Xy = np.array([0, ComienzoZonaNoAfectadaX, 0.5, FinZonaNoAfectadaX, 1])
            Yx = np.array([LimiteBajoY, ComienzoZonaNoAfectadaY, (self.Desplazamiento[1] if limiteAbajoDer is None else Desplazamiento[1]), FinZonaNoAfectadaY, LimiteAltoY])
            Yy = np.array([0, ComienzoZonaNoAfectadaY, 0.5, FinZonaNoAfectadaY, 1])

            # Crear la función polinómica
            polinomioX = np.poly1d(np.polyfit(Xx, Xy, 4))
            polinomioY = np.poly1d(np.polyfit(Yx, Yy, 4))

            # Calcular el valor ponderado
            return np.array([np.clip(polinomioX(mirada[0]), 0, 1), np.clip(polinomioY(mirada[1]), 0, 1)])

        def calcular_distancia(mirada, esquina):
            return np.sqrt((mirada[0] - esquina[0])**2 + (mirada[1] - esquina[1])**2)

        # Definir las cuatro esquinas
        esquinas = []
        if limiteAbajoDer is None:
            esquinas = [self.limiteAbajoIzq, self.limiteAbajoDer, self.limiteArribaIzq, self.limiteArribaDer]
        else:
            esquinas = [limiteAbajoIzq, limiteAbajoDer, limiteArribaIzq, limiteArribaDer]
        ponderaciones = []

        # Calcular la ponderación para cada esquina
        for esquina_limites in esquinas:
            ponderacion_esquina = ponderar_esquina(mirada, calcular_limites_esquina(esquinas.index(esquina_limites)))
            ponderaciones.append(ponderacion_esquina)

        
        # Calcular la distancia de la mirada a cada esquina
        distancias = [calcular_distancia(mirada, esquina) for esquina in esquinas]

        # Normalizar las distancias para obtener pesos de ponderación
        pesos = np.array([1 / (distancia*2 + 1) for distancia in distancias])  # Cambio aquí

        # Realizar normalizacion min-max de los pesos
        pesos = (pesos - np.min(pesos)) / (np.max(pesos) - np.min(pesos))

        # Sumar los pesos
        suma_pesos = np.sum(pesos)

        # Ponderar las ponderaciones de acuerdo a las distancias
        ponderacion_final = np.zeros_like(ponderaciones[0])
        for i, ponderacion_esquina in enumerate(ponderaciones):
            ponderacion_final += ponderacion_esquina * (pesos[i] / suma_pesos)

        return ponderacion_final

            
#--------------------------------FUNCIONES PARA LOS TABLEROS--------------------------------
#-------------------------------------------------------------------------------------------

    # def cargar_tableros(self):
    #     idioma = self.get_idioma()
    #     filename = f'./tableros/{idioma}.xlsx'
    #     if os.path.isfile(filename):
    #         xls = pd.ExcelFile(filename)
    #         for sheet_name in xls.sheet_names:
    #             df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    #             palabras = df.values.tolist()  # Convierte el DataFrame a una lista de listas
    #             self.tableros[sheet_name] = palabras  # Añade el tablero al diccionario

    def cargar_tableros(self):
        idioma = self.get_idioma()
        filename = f'./tableros/tableros_{idioma}.xlsx'
        if os.path.isfile(filename):
            wb = openpyxl.load_workbook(filename)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                palabras_con_imagenes = []
                for row in ws.iter_rows():
                    fila = []
                    for i in range(0, len(row), 2):
                        imagen_celda = row[i]
                        palabra_celda = row[i + 1]
                        fila.append((imagen_celda.value, palabra_celda.value))
                    palabras_con_imagenes.append(fila)
                self.tableros[sheet_name] = palabras_con_imagenes


    def obtener_tablero(self, nombre):
        return self.tableros.get(nombre.lower())
    
    def obtener_tablero_inicial(self):
        # Se coge el tablero con indice 0 
        return list(self.tableros.keys())[0]
    
    def get_frase(self):
        return self.frase
    
    def añadir_palabra(self, palabra):
        self.frase += palabra + ' '

    def borrar_palabra(self):
        self.frase = ' '.join(self.frase.rstrip().split(' ')[:-1]) + ' '
        self.contador_borrar += 1

    def borrar_todo(self):
        numero_palabras = len(self.frase.split(' '))
        self.contador_borrar += numero_palabras
        self.frase = ''

    def get_frase_bien(self):
        frase = self.frase.lower()
        if self.modelo_gemini is not None:
            prompt = "Recibo una frase con palabras en infinitivo y el idioma en el que está escrita(Español o gallego). Tu tarea es transformar la frase para que las palabras estén en la forma correcta y coherente entre sí siendo coherente con el idioma. Devuelve SOLAMENTE la frase corregida.\nEjemplo:\nEntrada: YO QUERER COMER CARNE\nRespuesta: Yo quiero comer carne\n\nFrase: " + frase + "\nIdioma: " + self.get_idioma()
            try:
                frase = self.modelo_gemini.generate_content(prompt).text
            except:
                pass
        return frase
        

    def reproducir_texto(self):
        #Empezar un hulo separado:
        def reproducir_texto_hilo():
            self.text_to_speech = win32com.client.Dispatch("SAPI.SpVoice").Speak(self.get_frase_bien())

        #Se crea un hilo para reproducir el texto
        self.tarea_hilo(lambda: reproducir_texto_hilo())

    def get_bloqueado(self):
        return self.bloqueado
    
    def set_bloqueado(self, valor):
        self.bloqueado = valor

    def cronometro(self, dt):
        self.cronometro_pruebas += dt

    def iniciar_cronometro(self):
        Clock.schedule_interval(self.cronometro, 0.01)
        self.contador_borrar = 0

    def get_cronometro(self):
        return self.cronometro_pruebas
    
    #Detiene el cronometro y guarda en el txt de resultados de los test el tiempo seguido de la frase
    def stop_cronometro(self):
        Clock.unschedule(self.cronometro)


    def reiniciar_cronometro(self):
        Clock.unschedule(self.cronometro)
        self.cronometro_pruebas = 0
        self.contador_borrar = 0


    
    #----------------------------------- FUNCIONES PARA EL REEENTRENAMIENTO --------------------------------
    #-----------------------------------------------------------------------------------------------------

    def reentrenar(self):
        # Se obtienen los datos
        self.input = np.array(self.input)
        self.output = np.array(self.output)

        # Si los datos estan vacios se pone el porcentaje a 100
        if len(self.input) < 10:
            print("No hay datos para reentrenar")
            self.porcentaje_reent = -1
            self.input = []
            self.output = []
            return

        # Se eliminan los datos con el ojo cerrado
        index = np.where(self.input[:, -2] < self.input[:, -1])
        self.input = np.delete(self.input, index, axis=0)
        self.output = np.delete(self.output, index, axis=0)

        # Se obtiene el conjunto 
        normalizar_funcion = getattr(Conjuntos, f'conjunto_{self.conjunto}')
        input_conj = normalizar_funcion(self.input)

        # Se convierten a tensores
        input_train = torch.from_numpy(input_conj).float()
        output_train = torch.from_numpy(self.output).float()

        # Se reentrena al modelo
        # Definir el optimizador
        optimizer = optim.Adam(self.modelo.parameters(), lr=0.001)

        train_losses = []
        best_loss = float('inf')
        models = []
        loss = nn.MSELoss()
        first_loss = loss(self.modelo(input_train), output_train).item()
        
        if self.optimizar:
            self.optimizando = True
            self.progreso_opt = 0

        #COMPROBAR EL NUMERO DE EPOCHS CREO QUE SON DEMASIADOS FAVORECE AL OVERFITTING
        for epoch in range(self.numero_epochs+1):
            optimizer.zero_grad()

            # Entrenamiento y cálculo de la pérdida
            train_predictions = self.modelo(input_train)
            train_loss = loss(train_predictions, output_train)
            train_losses.append(train_loss.item())

            # Guardar el mejor modelo
            if train_loss.item() < best_loss:
                best_loss = train_loss.item()
            models.append(self.modelo)

            # Actualizar el modelo
            train_loss.backward()
            optimizer.step()

            #Guardar el porcentaje de epoch que llevamos
            self.porcentaje_reent = int((epoch/self.numero_epochs)*100)

        self.modelo = models[train_losses.index(min(train_losses))]
        
        # Se optimizan las esquinas
        if self.optimizar:
            print("Optimizando las esquinas")
            indices, error_esquinas = self.optimizar_esquinas()

            #Calcular la perdida en las esquinas
            input_esq_sin = self.input[indices]
            output_esq_sin = self.output[indices]
            input_conj_esq_sin = normalizar_funcion(input_esq_sin)
            input_conj_esq_sin_tensor = torch.from_numpy(input_conj_esq_sin).float()
            train_predictions_esq_sin = self.modelo(input_conj_esq_sin_tensor)
            train_predictions_esq_sin_tensor = torch.tensor(train_predictions_esq_sin).float()
            output_esq_sin_tensor = torch.tensor(output_esq_sin).float()
            error_esquinas_sin = mse_loss(output_esq_sin_tensor, train_predictions_esq_sin_tensor).item()
            print("Error en las esquinas sin optimizar: ", error_esquinas_sin)
            print("Error en las esquinas optimizadas: ", error_esquinas)

            # Calcular la perdida en toda la pantalla
            input_conj = normalizar_funcion(self.input)
            input_conj_tensor = torch.from_numpy(input_conj).float()
            train_predictions = self.modelo(input_conj_tensor).detach().numpy()
            train_predictions_ponderadas = []
            for i, prediccion in enumerate(train_predictions):
                train_predictions_ponderadas.append(self.ponderar(prediccion))
            train_predictions_ponderadas_tensor = torch.tensor(train_predictions_ponderadas)
            perdida_final = mse_loss(output_train, train_predictions_ponderadas_tensor).item()
            print("Perdida final con las esquinas optimizadas: ", perdida_final)
        print("Perdida final sin las esquinas optimizadas: ", min(train_losses))
        print("Perdida antes del reentreno a los valores del usuario: ", first_loss)



    def descartar_reentrenamientos(self):
        if self.numero_entrenamientos == 0:
            self.mensaje(self.get_string("mensaje_no_reentrenamientos"))
            return
        self.modelo = torch.load(self.modelo_org)
        self.mensaje(self.get_string("mensaje_descartados"))
        self.numero_entrenamientos = 0

        # Limpiamos los datos de reentrenamiento
        self.input = []
        self.output = []

        # Reajustamos las esquinas a su valor orignial
        self.limiteAbajoIzq = self.limiteAbajoIzq_org
        self.limiteAbajoDer = self.limiteAbajoDer_org
        self.limiteArribaIzq = self.limiteArribaIzq_org
        self.limiteArribaDer = self.limiteArribaDer_org
        self.Desplazamiento = self.Desplazamiento_org


    def optimizar_esquinas(self):
        #Recoger los indices de los outputs que esten en las zonas (0-0,25;0-0,25), (0,75-1;0-0,25), (0-0,25;0,75-1), (0,75-1;0,75-1)
        indices = []
        for i, output in enumerate(self.output):
            if output[0] < 0.25 and output[1] < 0.25:
                indices.append(i)
            elif output[0] > 0.75 and output[1] < 0.25:
                indices.append(i)
            elif output[0] < 0.25 and output[1] > 0.75:
                indices.append(i)
            elif output[0] > 0.75 and output[1] > 0.75:
                indices.append(i)
        
        # Hacer un nuevo input y output con los indices
        input_opt = self.input[indices]
        output_opt = self.output[indices]

        # Se obtiene el conjunto y los tensores
        normalizar_funcion = getattr(Conjuntos, f'conjunto_{self.conjunto}')
        input_conj_opt = normalizar_funcion(input_opt)
        input_conj_opt = torch.from_numpy(input_conj_opt).float()

        def objective(trial):
            # Definir los límites de las esquinas y el desplazamiento
            limiteAbajoIzqX = trial.suggest_float('limiteAbajoIzqX', 0.00, 0.15)
            limiteAbajoIzqY = trial.suggest_float('limiteAbajoIzqY', 0.00, 0.15)
            limiteAbajoDerX = trial.suggest_float('limiteAbajoDerX', 0.85, 1.00)
            limiteAbajoDerY = trial.suggest_float('limiteAbajoDerY', 0.00, 0.15)
            limiteArribaIzqX = trial.suggest_float('limiteArribaIzqX', 0.00, 0.15)
            limiteArribaIzqY = trial.suggest_float('limiteArribaIzqY', 0.85, 1.00)
            limiteArribaDerX = trial.suggest_float('limiteArribaDerX', 0.85, 1.00)
            limiteArribaDerY = trial.suggest_float('limiteArribaDerY', 0.85, 1.00)
            DesplazamientoX = trial.suggest_float('DesplazamientoX', 0.46, 0.54)
            DesplazamientoY = trial.suggest_float('DesplazamientoY', 0.46, 0.54)

            limiteAbajoIzq = [limiteAbajoIzqX, limiteAbajoIzqY]
            limiteAbajoDer = [limiteAbajoDerX, limiteAbajoDerY]
            limiteArribaIzq = [limiteArribaIzqX, limiteArribaIzqY]
            limiteArribaDer = [limiteArribaDerX, limiteArribaDerY]
            Desplazamiento = [DesplazamientoX, DesplazamientoY]

            # Calcular las predicciones del modelo
            predicciones = self.modelo((input_conj_opt)).detach().numpy()

            for i, prediccion in enumerate(predicciones):
                predicciones[i] = self.ponderar(prediccion, limiteAbajoIzq, limiteAbajoDer, limiteArribaIzq, limiteArribaDer, Desplazamiento)

            
            # Calcular el error
            output_tensor = torch.from_numpy(output_opt).float()
            predicciones_tensor = torch.from_numpy(predicciones).float()

            error = mse_loss(output_tensor, predicciones_tensor).item()

            return error
        
        
        def actualizar_progreso_opt(study, trial):
            self.progreso_opt = int(((trial.number+1) / self.trials_opt) * 100)

        study = optuna.create_study(direction='minimize')
        study.optimize(objective, n_trials=self.trials_opt, callbacks=[actualizar_progreso_opt],show_progress_bar=False)

        # Guardar los resultados
        self.limiteAbajoIzq = [study.best_params['limiteAbajoIzqX'], study.best_params['limiteAbajoIzqY']]
        self.limiteAbajoDer = [study.best_params['limiteAbajoDerX'], study.best_params['limiteAbajoDerY']]
        self.limiteArribaIzq = [study.best_params['limiteArribaIzqX'], study.best_params['limiteArribaIzqY']]
        self.limiteArribaDer = [study.best_params['limiteArribaDerX'], study.best_params['limiteArribaDerY']]
        self.Desplazamiento = [study.best_params['DesplazamientoX'], study.best_params['DesplazamientoY']]
        
        self.optimizando = False
        self.progreso_opt = 0
        return indices, study.best_value  